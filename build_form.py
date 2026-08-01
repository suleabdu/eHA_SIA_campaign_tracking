#!/usr/bin/env python3
"""
Builds HH2026_v1.xlsx - the XLSForm for the Bansara Integrated Child Health
and Antimicrobial Resistance Survey 2026 household questionnaire (Form HH/2026/v1).

Run: python3 build_form.py
Then convert with: pyxform HH2026_v1.xlsx HH2026_v1.xml   (see convert.sh)
"""
import openpyxl

HEADER = [
    "type", "name",
    "label::English (en)", "label::Hausa (ha)",
    "hint::English (en)",
    "appearance", "default", "calculation",
    "constraint", "constraint_message::English (en)",
    "relevant", "required", "required_message::English (en)",
    "read_only", "choice_filter", "repeat_count", "parameters", "body::accuracyThreshold",
]
COLIDX = {c: i for i, c in enumerate(HEADER)}

def row(type_, name="", en="", ha="", hint="", appearance="", default="", calc="",
        constraint="", cmsg="", relevant="", required="", reqmsg="",
        readonly="", choice_filter="", repeat_count="", parameters="", accuracy=""):
    r = [""] * len(HEADER)
    r[COLIDX["type"]] = type_
    r[COLIDX["name"]] = name
    r[COLIDX["label::English (en)"]] = en
    r[COLIDX["label::Hausa (ha)"]] = ha
    r[COLIDX["hint::English (en)"]] = hint
    r[COLIDX["appearance"]] = appearance
    r[COLIDX["default"]] = default
    r[COLIDX["calculation"]] = calc
    r[COLIDX["constraint"]] = constraint
    r[COLIDX["constraint_message::English (en)"]] = cmsg
    r[COLIDX["relevant"]] = relevant
    r[COLIDX["required"]] = required
    r[COLIDX["required_message::English (en)"]] = reqmsg
    r[COLIDX["read_only"]] = readonly
    r[COLIDX["choice_filter"]] = choice_filter
    r[COLIDX["repeat_count"]] = repeat_count
    r[COLIDX["parameters"]] = parameters
    r[COLIDX["body::accuracyThreshold"]] = accuracy
    return r

S = []  # survey rows

# ---------------------------------------------------------------------------
# METADATA (hidden, device/session provenance - used later for fabrication
# detection and version tracking; not shown to the respondent)
# ---------------------------------------------------------------------------
S.append(row("start", "start_ts"))
S.append(row("end", "end_ts"))
S.append(row("today", "today_date"))
S.append(row("deviceid", "device_id"))
S.append(row("username", "device_username"))
S.append(row("calculate", "form_version_tag", calc="'HH2026v1'"))

S.append(row("note", "confidentiality_notice",
             "CONFIDENTIALITY: The information recorded on this form is confidential. It identifies individual households and children and includes biological specimen identifiers. It may be used only for the purposes of this survey and must not be shown or disclosed to anyone who is not a member of the survey team. Ethics approval: BSHREC/2026/041.",
             "SIRRI: Bayanan da ake rubuta a wannan fam sirri ne. Ba za a nuna ko fadin su ga wanda ba memba na tawagar bincike ba. Amincewar da'a: BSHREC/2026/041."))

# ---------------------------------------------------------------------------
# SECTION 1: Household identification
# ---------------------------------------------------------------------------
S.append(row("begin group", "sec1", "Section 1: Household identification",
             "Kashi na 1: Bayanan gida"))

S.append(row("calculate", "state_name", calc="'Bansara'"))

S.append(row("select_one_from_file lgas.csv", "lga",
             "1.02 Local Government Area (LGA)", "1.02 Karamar Hukuma (LGA)"))

S.append(row("select_one_from_file wards.csv", "ward",
             "1.03 Ward", "1.03 Unguwa (Ward)",
             choice_filter="lga_code=${lga}"))

S.append(row("select_one_from_file settlements.csv", "settlement",
             "1.04 Settlement", "1.04 Sulale/Ƴari (Settlement)",
             appearance="autocomplete",
             choice_filter="ward_code=${ward}",
             hint="Type to search within this ward's settlement list."))

S.append(row("select_one yesno", "settlement_local_name_diff",
             "1.05 Is the settlement known locally by a different name?",
             "1.05 Ana kiran wurin da suna daban a gida?"))
S.append(row("text", "settlement_local_name", "Local name used",
             "Sunan da ake amfani da shi a gida",
             relevant="${settlement_local_name_diff}='1'",
             required="yes"))

S.append(row("integer", "structure_number", "1.06 Structure number painted on the dwelling",
             "1.06 Lambar gini",
             constraint=". >= 1 and . <= 999",
             cmsg="Structure number must be between 1 and 999 (3-digit field on the paper form).",
             required="yes"))

S.append(row("integer", "hh_serial", "1.07 Household serial number within the settlement",
             "1.07 Lambar gida a cikin sulale",
             constraint=". >= 1 and . <= 999",
             cmsg="Serial number must be between 1 and 999.",
             required="yes"))

# Enumerator identity: pulled from the device's provisioned username rather than
# re-keyed, then cross-checked against the staff roster. See constraint register
# item C-08.
S.append(row("calculate", "enumerator_code", calc="${device_username}"))
S.append(row("calculate", "enumerator_code_valid",
             calc="if(string-length(pulldata('staff_roster','role','name',${enumerator_code}))>0,1,0)"))
S.append(row("note", "enumerator_code_warn",
             "WARNING: device username '${enumerator_code}' was not found in the staff roster. Do not proceed - contact your supervisor before continuing; this device may be mis-provisioned.",
             relevant="${enumerator_code_valid}=0"))
S.append(row("calculate", "enumerator_name",
             calc="pulldata('staff_roster','label','name',${enumerator_code})"))
S.append(row("calculate", "enumerator_team",
             calc="pulldata('staff_roster','team_code','name',${enumerator_code})"))
S.append(row("note", "enumerator_confirm_note",
             "Logged in as ${enumerator_name}, team ${enumerator_team}.",
             relevant="${enumerator_code_valid}=1"))

S.append(row("date", "visit_date", "1.10 Date of visit", "1.10 Ranar ziyara",
             constraint=". >= date('2026-06-01') and . <= date('2026-06-30')",
             cmsg="Date of visit must fall within the fieldwork period, 1-30 June 2026.",
             required="yes"))

S.append(row("geopoint", "gps_dwelling",
             "1.11 Record the GPS reading taken at the entrance to the dwelling.",
             "1.11 Dauki GPS a kofar gidan",
             required="yes", accuracy="15"))

S.append(row("select_one yesnodk", "visited_oct2025",
             "1.12 Was this household visited during the October 2025 round?",
             "1.12 An ziyarci wannan gidan a zagayen Oktoba 2025?",
             required="yes"))

S.append(row("text", "prior_hh_id",
             "1.13 Record the household identifier allocated in the October 2025 round (format BAN-######).",
             "1.13 Rubuta lambar gidan da aka bayar a zagayen Oktoba 2025",
             relevant="${visited_oct2025}='1'",
             required="yes",
             constraint="regex(.,'^BAN-[0-9]{6}$') and pulldata('previous_round_households','settlement_id','household_id',.)=${settlement}",
             cmsg="Identifier must be in the form BAN-000000 and must match a household previously recorded in this same settlement. If the household moved settlement since October 2025, code 1.12 as 'Do not know' and escalate to your supervisor rather than forcing a match."))

S.append(row("select_one result_visit", "result_of_visit",
             "1.14 Result of visit", "1.14 Sakamakon ziyara",
             required="yes"))

S.append(row("end group"))

# Everything below Section 1 only applies if the visit produced a completed
# interview (paper: "2, 3 or 4 -> END" at 1.14).
S.append(row("begin group", "sec_gate1", "", relevant="${result_of_visit}='1'"))

# ---------------------------------------------------------------------------
# SECTION 2: Consent
# ---------------------------------------------------------------------------
S.append(row("begin group", "sec2", "Section 2: Consent", "Kashi na 2: Yarda"))

S.append(row("select_one yesno", "consent_read",
             "2.01 Consent statement read aloud to the respondent in full?",
             "2.01 An karanta bayanin yarda gaba daya ga wanda ake tambaya?",
             required="yes"))

S.append(row("select_one consent", "consent_given",
             "2.02 Does the respondent consent to the household interview?",
             "2.02 Wanda ake tambaya ya yarda da tambayoyin?",
             required="yes"))

S.append(row("select_one relationship_head", "respondent_relationship",
             "2.03 Relationship of the respondent to the head of household",
             "2.03 Alakar wanda ake tambaya da shugaban gida",
             relevant="${consent_given}='1'",
             required="yes"))

S.append(row("end group"))

# Everything below Section 2 only applies if consent was given.
S.append(row("begin group", "sec_gate2", "", relevant="${consent_given}='1'"))

# ---------------------------------------------------------------------------
# SECTION 3: Household roster
# ---------------------------------------------------------------------------
S.append(row("begin group", "sec3", "Section 3: Household roster", "Kashi na 3: Jerin mazauna gida"))

S.append(row("integer", "hh_size_stated",
             "3.01 How many people usually live in this household?",
             "3.01 Mutane nawa ne suke zaune a gidan nan?",
             constraint=". >= 1 and . <= 30",
             cmsg="Enter a household size between 1 and 30. If genuinely larger, enter 30 and describe in the notes at 7.02 - do not use a non-response code here (see constraint register C-01).",
             required="yes"))

S.append(row("begin repeat", "roster", "Household member", "Mazaunin gida"))

S.append(row("text", "r_name", "(2) Name or initials", "(2) Suna ko gajerun haruffa",
             required="yes"))

S.append(row("select_one relationship_head", "r_relationship",
             "(3) Relationship to head", "(3) Alaka da shugaban gida",
             required="yes"))

S.append(row("select_one sex", "r_sex", "(4) Sex", "(4) Jinsi", required="yes"))

S.append(row("select_one yesno", "r_is_under5",
             "Is this person under five years old?",
             "Wannan mutum yana kasa da shekara biyar?",
             hint="Interviewer instruction (not printed on the paper roster as a question): this determines whether age is recorded in months or years, per the instruction above the roster grid.",
             required="yes"))

S.append(row("select_one yesno", "r_age_unknown",
             "Age not known (respondent cannot state an age)?",
             "Ba a san shekarun mutumin ba?",
             required="yes"))

S.append(row("integer", "r_age_years", "(5) Age in years",
             "(5) Shekaru",
             relevant="${r_is_under5}='2' and ${r_age_unknown}='2'",
             constraint=". >= 5 and . <= 97",
             cmsg="Age in years must be 5-97. Top-code any person aged 97 or older as 97 (see constraint register C-02) rather than using a sentinel value.",
             required="${r_is_under5}='2' and ${r_age_unknown}='2'"))

S.append(row("integer", "r_age_months", "(6) Age in months (under 5 only)",
             "(6) Watanni (kasa da shekara 5 kawai)",
             relevant="${r_is_under5}='1' and ${r_age_unknown}='2'",
             constraint=". >= 0 and . <= 59",
             cmsg="Age in completed months must be 0-59 for a child recorded as under five.",
             required="${r_is_under5}='1' and ${r_age_unknown}='2'"))

S.append(row("calculate", "r_eligible",
             calc="if(${r_is_under5}='1' and ${r_age_unknown}='2' and ${r_age_months}>=9 and ${r_age_months}<=59,1,0)"))

S.append(row("end repeat", "roster"))

S.append(row("calculate", "roster_person_count", calc="count(${roster})"))
S.append(row("calculate", "hh_size_mismatch",
             calc="if(${hh_size_stated}!=${roster_person_count},1,0)"))
S.append(row("text", "hh_size_discrepancy_note",
             "3.01 states ${hh_size_stated} usual residents, but ${roster_person_count} were listed in the roster above. Explain the discrepancy (e.g. member away at time of listing, late addition) before continuing.",
             "Akwai bambanci tsakanin 3.01 da jerin mazauna. Bayyana dalili.",
             relevant="${hh_size_mismatch}=1",
             required="yes"))

S.append(row("integer", "eligible_count_stated",
             "3.02 From column (7), how many children in this household are aged 9 to 59 completed months?",
             "3.02 Yara nawa ne masu shekara 9 zuwa 59 a wata a gidan nan?",
             constraint=". >= 0 and . <= 15",
             cmsg="Enter the number of eligible children, 0-15.",
             required="yes"))

S.append(row("calculate", "roster_eligible_count", calc="count(${roster}[r_eligible=1])"))
S.append(row("calculate", "eligible_mismatch",
             calc="if(${eligible_count_stated}!=${roster_eligible_count},1,0)"))
S.append(row("text", "eligible_discrepancy_note",
             "3.02 states ${eligible_count_stated} eligible children, but ${roster_eligible_count} roster entries meet the 9-59 month criterion. Check the ages recorded above and explain any remaining discrepancy.",
             "Akwai bambanci tsakanin 3.02 da adadin da aka lissafa. Bayyana dalili.",
             relevant="${eligible_mismatch}=1",
             required="yes"))

S.append(row("end group"))

# ---------------------------------------------------------------------------
# SECTION 4 + 5 (combined per child, matching the paper's "same page" rule):
# Child module and specimen collection. One repeat instance per eligible
# child. The enumerator keys only the roster line number; name/age/sex are
# pulled automatically via indexed-repeat() rather than re-copied by hand.
# ---------------------------------------------------------------------------
S.append(row("begin group", "sec45", "Sections 4 and 5: Child module and specimen collection",
             relevant="${roster_eligible_count} > 0 or ${eligible_count_stated} > 0"))

S.append(row("note", "child_module_intro",
             "There are ${roster_eligible_count} eligible children in the roster (9-59 completed months). Add one child module for each, in any order.",
             "Akwai yara ${roster_eligible_count} da suka cancanta a jerin mazauna. Kara shafi daya ga kowanne."))

S.append(row("begin repeat", "child_module", "Child module", "Shafin yaro"))

S.append(row("integer", "c_line_number",
             "4.01 Line number of this child in the Section 3 roster",
             "4.01 Lambar layi na yaro a jerin mazauna",
             constraint=(
                 "indexed-repeat(${r_is_under5},${roster},.)='1' and "
                 "indexed-repeat(${r_age_unknown},${roster},.)='2' and "
                 "indexed-repeat(${r_age_months},${roster},.)>=9 and "
                 "indexed-repeat(${r_age_months},${roster},.)<=59"
             ),
             cmsg="This line number does not correspond to a roster entry aged 9-59 completed months. Check column (7) of the roster.",
             required="yes"))

S.append(row("calculate", "c_name", calc="indexed-repeat(${r_name},${roster},${c_line_number})"))
S.append(row("calculate", "c_age_months", calc="indexed-repeat(${r_age_months},${roster},${c_line_number})"))
S.append(row("calculate", "c_sex", calc="indexed-repeat(${r_sex},${roster},${c_line_number})"))
S.append(row("note", "c_identity_note",
             "4.02-4.04 Child: ${c_name}, ${c_age_months} months, sex code ${c_sex} (pulled automatically from the roster - verify this is the correct child before continuing)."))

S.append(row("select_one yesno", "c_weight_not_measured", "Weight not measured?",
             "Ba a auna nauyi ba?", required="yes"))
S.append(row("decimal", "c_weight_kg", "4.05 Weight of the child (kg)", "4.05 Nauyin yaro (kg)",
             relevant="${c_weight_not_measured}='2'",
             constraint=". >= 2.0 and . <= 30.0",
             cmsg="Weight out of plausible range for 9-59 months (2.0-30.0 kg; WHO Child Growth Standards range with a data-entry buffer - judgement threshold, see C-05).",
             required="${c_weight_not_measured}='2'"))

S.append(row("select_one yesno", "c_height_not_measured", "Length/height not measured?",
             "Ba a auna tsawo ba?", required="yes"))
S.append(row("decimal", "c_height_cm", "4.06 Length or height of the child (cm)", "4.06 Tsawon yaro (cm)",
             relevant="${c_height_not_measured}='2'",
             constraint=". >= 45 and . <= 120",
             cmsg="Length/height out of plausible range for 9-59 months (45-120 cm - judgement threshold, see C-05).",
             required="${c_height_not_measured}='2'"))

S.append(row("select_one measure_position", "c_measure_position",
             "4.07 Position in which the child was measured", "4.07 Yadda aka auna yaro",
             relevant="${c_height_not_measured}='2'",
             required="${c_height_not_measured}='2'"))

S.append(row("calculate", "c_position_expected",
             calc="if(${c_age_months}<24,'1','2')"))
S.append(row("calculate", "c_position_mismatch",
             calc="if(${c_height_not_measured}='2' and ${c_measure_position}!=${c_position_expected},1,0)"))
S.append(row("text", "c_position_mismatch_note",
             "Standard practice is recumbent length under 24 months and standing height at 24 months or older; this child was measured the other way. Explain (e.g. child would not lie still, could not stand unaided).",
             "An auna yaro ta hanyar da ba a saba amfani da ita ba don shekarunsa. Bayyana dalili.",
             relevant="${c_position_mismatch}=1",
             required="yes"))

S.append(row("select_one card_seen", "c_card_seen",
             "4.08 May I see the child's vaccination card or health record?",
             "4.08 Zan iya ganin katin allurar rigakafi na yaro?",
             required="yes"))

S.append(row("select_one yesno", "c_measles_recorded",
             "4.09 Copy from the card: is a measles dose recorded?",
             "4.09 An rubuta maganin kyanda a kati?",
             relevant="${c_card_seen}='1'", required="yes"))

S.append(row("select_one yesnodk", "c_measles_ever",
             "4.10 Has this child ever received a measles vaccination?",
             "4.10 Yaro ya taba samun maganin kyanda?",
             relevant="${c_card_seen}='2'", required="yes"))

S.append(row("select_one yesnodk", "c_diarrhoea_14d",
             "4.11 Has this child had diarrhoea in the past 14 days?",
             "4.11 Yaro ya sha zawo a kwanaki 14 da suka wuce?",
             required="yes"))

S.append(row("select_one yesnodk", "c_antibiotic_30d",
             "4.12 Has this child taken any antibiotic medicine in the past 30 days?",
             "4.12 Yaro ya sha wani maganin rigakafin kwayoyin cuta a kwanaki 30 da suka wuce?",
             required="yes"))

S.append(row("select_one_from_file antibiotics_PLACEHOLDER.csv", "c_antibiotic_which",
             "4.13 Which antibiotic was taken?", "4.13 Wane magani ne aka sha?",
             relevant="${c_antibiotic_30d}='1'",
             required="${c_antibiotic_30d}='1'",
             hint="PLACEHOLDER LIST - see constraint register C-11. Not the ministry's authoritative medicine list."))

S.append(row("text", "c_antibiotic_which_other", "4.14 If code 96, name of the medicine as reported",
             "4.14 Idan lamba 96, rubuta sunan maganin",
             relevant="${c_antibiotic_which}='96'", required="yes"))

S.append(row("select_one yesnodk", "c_antibiotic_no_prescription",
             "4.15 Was the medicine obtained without a prescription from a health worker?",
             "4.15 An sami maganin ba tare da takardar likita ba?",
             relevant="${c_antibiotic_30d}='1'", required="yes"))

S.append(row("select_one photo_status", "c_antibiotic_photo",
             "4.16 Was a photograph of the medicine packaging taken?",
             "4.16 An dauki hoton fakitin maganin?",
             relevant="${c_antibiotic_30d}='1'", required="yes"))

# --- Section 5: specimen collection (same page/repeat instance as Section 4) ---
S.append(row("calculate", "c_age12plus", calc="if(${c_age_months}>=12,1,0)"))
# 5.01 is derived from the age already captured at 4.03 rather than re-asked -
# see constraint register, defect D-04 (internal contradiction risk removed).

S.append(row("select_one yesno", "c_specimen_obtained",
             "5.02 Was a stool specimen obtained from this child?",
             "5.02 An samu samfurin kashi daga yaro?",
             relevant="${c_age12plus}=1", required="yes"))

S.append(row("integer", "c_specimen_label_number",
             "5.03 Specimen label number (6 digits, before the check digit)",
             "5.03 Lambar alama (adadi 6)",
             relevant="${c_specimen_obtained}='1'",
             calc="", required="${c_specimen_obtained}='1'"))

S.append(row("calculate", "c_label_range_start",
             calc="pulldata('specimen_label_allocation','range_start','team_code',${enumerator_team})"))
S.append(row("calculate", "c_label_range_end",
             calc="pulldata('specimen_label_allocation','range_end','team_code',${enumerator_team})"))

S.append(row("text", "c_specimen_check_digit",
             "5.03 Check digit (0-9 or X)", "5.03 Lambar tabbatarwa",
             relevant="${c_specimen_obtained}='1'",
             required="${c_specimen_obtained}='1'"))

# Check-digit computation: modulus 11, weights 2-7 applied right to left over
# the 6-digit label number; remainder 10 recorded as X (per
# specimen_label_allocation.csv). Built from substr() because JavaRosa/XPath
# 1.0 has no loop construct - see constraint register C-09 and check_digit.py
# for the reference implementation and test cases.
S.append(row("calculate", "cd_str", calc="string(${c_specimen_label_number})"))
S.append(row("calculate", "cd_d1", calc="number(substr(${cd_str},0,1))"))
S.append(row("calculate", "cd_d2", calc="number(substr(${cd_str},1,2))"))
S.append(row("calculate", "cd_d3", calc="number(substr(${cd_str},2,3))"))
S.append(row("calculate", "cd_d4", calc="number(substr(${cd_str},3,4))"))
S.append(row("calculate", "cd_d5", calc="number(substr(${cd_str},4,5))"))
S.append(row("calculate", "cd_d6", calc="number(substr(${cd_str},5,6))"))
S.append(row("calculate", "cd_weighted_sum",
             calc="(${cd_d6}*2)+(${cd_d5}*3)+(${cd_d4}*4)+(${cd_d3}*5)+(${cd_d2}*6)+(${cd_d1}*7)"))
S.append(row("calculate", "cd_remainder", calc="${cd_weighted_sum} mod 11"))
S.append(row("calculate", "cd_expected",
             calc="if(${cd_remainder}=10,'X',string(${cd_remainder}))"))

S.append(row("calculate", "c_specimen_label_valid",
             calc=(
                 "if(${c_specimen_label_number}>=${c_label_range_start} and "
                 "${c_specimen_label_number}<=${c_label_range_end} and "
                 "upper-case(${c_specimen_check_digit})=${cd_expected},1,0)"
             )))
S.append(row("note", "c_specimen_label_warn",
             "This label number/check digit combination is not valid for this device's team allocation. Re-check the pre-printed label before typing an override justification below.",
             "Lambar alama da lambar tabbatarwa ba daidai suke ba. Duba alamar da kyau.",
             relevant="${c_specimen_obtained}='1' and ${c_specimen_label_valid}=0"))
S.append(row("text", "c_specimen_label_override_note",
             "Label/check digit did not validate against this team's allocated range. Re-checked the physical label and confirm the value entered is exactly what is printed, or explain (e.g. label from another team's spare stock was used).",
             "An sake duba alamar; tabbatar da lambar da aka shigar daidai take da abin da yake a alamar, ko bayyana dalili.",
             relevant="${c_specimen_obtained}='1' and ${c_specimen_label_valid}=0",
             required="yes"))

S.append(row("calculate", "c_specimen_label_full",
             calc="if(${c_specimen_obtained}='1',concat('BSN',${cd_str},'-',upper-case(${c_specimen_check_digit})),'')"))

S.append(row("time", "c_specimen_time_coldbox",
             "5.04 Time the specimen was placed in the cold box", "5.04 Lokacin da aka sanya samfurin a akwatin sanyi",
             relevant="${c_specimen_obtained}='1'", required="yes"))

S.append(row("decimal", "c_specimen_temp",
             "5.05 Temperature shown on the cold box thermometer at that time",
             "5.05 Zafin da ma'aunin zafi ya nuna",
             relevant="${c_specimen_obtained}='1'",
             constraint=". >= -1 and . <= 10",
             cmsg="Cold box temperature out of the plausible transport range (-1 to 10 degrees C - WHO EPI cold-chain reference band widened for data-entry error catching, judgement threshold, see C-05). Values outside 2-8C are valid readings and should still be recorded; this only catches implausible entries like 28.0 for 2.8.",
             required="yes"))

S.append(row("select_one specimen_no_reason", "c_specimen_no_reason",
             "5.06 Reason no specimen was obtained", "5.06 Dalilin da ba a samu samfuri ba",
             relevant="${c_age12plus}=1 and ${c_specimen_obtained}='2'", required="yes"))

S.append(row("text", "c_specimen_no_other", "5.07 If code 96, specify",
             "5.07 Idan lamba 96, bayyana",
             relevant="${c_specimen_no_reason}='96'", required="yes"))

S.append(row("end repeat", "child_module"))

S.append(row("calculate", "completed_child_modules", calc="count(${child_module})"))
S.append(row("calculate", "child_count_mismatch",
             calc="if(${roster_eligible_count}!=${completed_child_modules},1,0)"))
S.append(row("text", "child_count_discrepancy_note",
             "The roster identifies ${roster_eligible_count} eligible children but ${completed_child_modules} child modules were completed. Explain (e.g. child absent, caregiver ended interview early) before continuing.",
             "Akwai bambanci tsakanin adadin yara da suka cancanta da adadin shafukan da aka cika. Bayyana dalili.",
             relevant="${child_count_mismatch}=1",
             required="yes"))

S.append(row("end group"))

# ---------------------------------------------------------------------------
# SECTION 6: Household environment
# ---------------------------------------------------------------------------
S.append(row("begin group", "sec6", "Section 6: Household environment", "Kashi na 6: Muhallin gida"))

S.append(row("select_one water_source", "water_source",
             "6.01 What is the main source of drinking water for members of this household?",
             "6.01 Ina ake samun ruwan sha na gidan nan?",
             required="yes"))

S.append(row("select_one toilet_facility", "toilet_facility",
             "6.02 What kind of toilet facility do members of this household usually use?",
             "6.02 Wace irin bayan gida ake amfani da ita?",
             required="yes"))

S.append(row("select_one yesno", "livestock_present",
             "6.03 Does this household keep poultry or livestock inside the compound?",
             "6.03 Gidan yana da kaji ko dabbobi a cikin gida?",
             required="yes"))

S.append(row("select_one yesnodk", "livestock_antibiotics",
             "6.04 Have any antibiotic medicines been given to these animals in the past 12 months?",
             "6.04 An ba dabbobin nan magungunan rigakafin kwayoyin cuta a cikin watanni 12 da suka wuce?",
             relevant="${livestock_present}='1'", required="yes"))

S.append(row("select_one handwash", "handwashing",
             "6.05 Observe: is there a handwashing station with both soap and water available?",
             "6.05 Duba: akwai wurin wanke hannu da sabulu da ruwa?",
             required="yes"))

S.append(row("select_one yesnodk", "diarrhoea_hh",
             "6.06 Has any member of this household had diarrhoea in the past two weeks?",
             "6.06 Wani a gidan ya sha zawo a makonni biyu da suka wuce?",
             required="yes"))

S.append(row("select_multiple possessions", "possessions",
             "6.07 Which of the following does this household own?",
             "6.07 Wanne daga cikin wadannan gidan yana da su?",
             required="yes"))

S.append(row("end group"))

# ---------------------------------------------------------------------------
# SECTION 7: Close-out and supervisor review
# ---------------------------------------------------------------------------
S.append(row("begin group", "sec7", "Section 7: Close-out and supervisor review",
             "Kashi na 7: Kammalawa da bitar mai kula"))

S.append(row("time", "interview_end_time", "7.01 Time the interview ended",
             "7.01 Lokacin da hirar ta kare", required="yes"))

S.append(row("calculate", "interview_duration_min",
             calc="(${end_ts} - ${start_ts}) div 60000"))

S.append(row("text", "observation_notes",
             "7.02 Record any observation that may help the office interpret this form.",
             "7.02 Rubuta duk wata lura da za ta taimaka wa ofis"))

S.append(row("select_one yesno", "enumerator_confirms",
             "7.03 I confirm the information on this form was recorded as stated by the respondent and/or observed by me.",
             "7.03 Na tabbatar bayanan sun dace da abin da aka fada/aka gani",
             required="yes",
             constraint=".='1'",
             cmsg="The enumerator must confirm before the form can be submitted. See constraint register, item 14 (digital signature substitute)."))

S.append(row("end group"))  # sec7

# Supervisor review (7.04-7.06). Implemented in the same form/device rather
# than as a separate ODK Central review workflow - see constraint register,
# item 14, for the alternative architecture considered and why it was not
# chosen for this exercise.
S.append(row("begin group", "sec7_supervisor", "Supervisor review", "Bitar mai kula"))

S.append(row("text", "supervisor_code", "7.04 Supervisor code",
             "7.04 Lambar mai kula",
             required="yes",
             constraint="string-length(pulldata('staff_roster','role','name',.))>0 and pulldata('staff_roster','role','name',.)='Team supervisor'",
             cmsg="Code not found in the staff roster as a team supervisor."))

S.append(row("select_one supervisor_decision", "supervisor_decision",
             "7.05 Supervisor decision on this form", "7.05 Shawarar mai kula",
             required="yes"))

S.append(row("text", "supervisor_return_reason",
             "State the reason for return or voiding.", "Bayyana dalilin mayarwa ko sokewa",
             relevant="${supervisor_decision}='2' or ${supervisor_decision}='3'",
             required="yes"))

S.append(row("calculate", "supervisor_review_date", calc="${today_date}"))

S.append(row("end group"))  # sec7_supervisor

S.append(row("end group"))  # sec_gate2
S.append(row("end group"))  # sec_gate1

S.append(row("calculate", "office_note",
             calc="'Section 8 (office use: receipt date, data-entry clerk code, second-entry verification) is a paper double-entry workflow artefact and has no digital equivalent; deliberately not carried into this form - see item 14.'"))

# ---------------------------------------------------------------------------
# CHOICES
# ---------------------------------------------------------------------------
CH_HEADER = ["list_name", "name", "label::English (en)", "label::Hausa (ha)"]
C = []
def choice(list_name, name, en, ha=""):
    C.append([list_name, name, en, ha])

for n, en, ha in [("1", "Yes", "Ee"), ("2", "No", "A'a")]:
    choice("yesno", n, en, ha)
for n, en, ha in [("1", "Yes", "Ee"), ("2", "No", "A'a"), ("8", "Do not know", "Ban sani ba")]:
    choice("yesnodk", n, en, ha)
for n, en in [("1", "Completed"), ("2", "Refused"), ("3", "No competent adult after three visits"),
              ("4", "Dwelling vacant or demolished")]:
    choice("result_visit", n, en)
for n, en in [("1", "Consent given"), ("2", "Consent refused")]:
    choice("consent", n, en)
for n, en in [("1", "Head"), ("2", "Spouse"), ("3", "Son or daughter"), ("4", "Parent"),
              ("5", "Other relative"), ("6", "Not related")]:
    choice("relationship_head", n, en)
for n, en in [("1", "Male"), ("2", "Female")]:
    choice("sex", n, en)
for n, en in [("1", "Recumbent length"), ("2", "Standing height")]:
    choice("measure_position", n, en)
for n, en in [("1", "Card seen"), ("2", "No card seen")]:
    choice("card_seen", n, en)
for n, en in [("1", "Caregiver refused"), ("2", "Child absent"), ("3", "Unable to produce"),
              ("4", "Container spoiled"), ("96", "Other")]:
    choice("specimen_no_reason", n, en)
for n, en in [("1", "Piped into dwelling"), ("2", "Piped into compound"), ("3", "Public tap or standpipe"),
              ("4", "Tube well or borehole"), ("5", "Protected dug well"), ("6", "Unprotected dug well"),
              ("7", "Protected spring"), ("8", "Unprotected spring"), ("9", "Rainwater"),
              ("10", "Tanker or cart"), ("11", "Surface water"),
              ("98", "Do not know / no response (added - see constraint register C-10)")]:
    choice("water_source", n, en)
for n, en in [("1", "Flush to sewer"), ("2", "Flush to septic tank"), ("3", "Flush to pit latrine"),
              ("4", "Ventilated improved pit"), ("5", "Pit latrine with slab"), ("6", "Pit latrine without slab"),
              ("7", "Composting toilet"), ("8", "Bucket"), ("9", "No facility or bush"),
              ("98", "Do not know / no response (added - see constraint register C-10)")]:
    choice("toilet_facility", n, en)
for n, en in [("1", "Observed, soap and water"), ("2", "Reported only, not observed"), ("3", "Not present")]:
    choice("handwash", n, en)
for n, en in [("A", "Radio"), ("B", "Television"), ("C", "Mobile telephone"), ("D", "Bicycle"),
              ("E", "Motorcycle"), ("F", "Car or truck"), ("G", "Refrigerator"), ("H", "None of these")]:
    choice("possessions", n, en)
for n, en in [("1", "Yes"), ("2", "No, not available"), ("3", "Caregiver declined")]:
    choice("photo_status", n, en)
for n, en in [("1", "Accept"), ("2", "Return for correction"), ("3", "Void")]:
    choice("supervisor_decision", n, en)

# ---------------------------------------------------------------------------
# SETTINGS
# ---------------------------------------------------------------------------
SETTINGS_HEADER = ["form_title", "form_id", "version", "default_language", "style"]
SETTINGS_ROW = ["Integrated Child Health and AMR Survey 2026 - Household Questionnaire",
                "hh2026_v1", "2026060100", "Hausa (ha)", "pages"]

# ---------------------------------------------------------------------------
# WRITE WORKBOOK
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "survey"
ws.append(HEADER)
for r in S:
    ws.append(r)

ws2 = wb.create_sheet("choices")
ws2.append(CH_HEADER)
for r in C:
    ws2.append(r)

ws3 = wb.create_sheet("settings")
ws3.append(SETTINGS_HEADER)
ws3.append(SETTINGS_ROW)

wb.save("HH2026_v1.xlsx")
print(f"Wrote HH2026_v1.xlsx: {len(S)} survey rows, {len(C)} choice rows.")
